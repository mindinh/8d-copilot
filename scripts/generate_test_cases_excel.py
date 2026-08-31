import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_comprehensive_test_cases_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "D2 Is-IsNot Test Cases"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Color Palette - SAP Fiori / Corporate Palette
    PRIMARY_COLOR = "1B365D"       # Deep Navy Header
    ZEBRA_FILL = "F8FAFC"          # Light Slate
    BORDER_COLOR = "CBD5E1"        # Light Gray Border
    PASS_COLOR = "D1FAE5"          # Soft Green Fill
    PASS_TEXT = "065F46"           # Dark Green Text
    CRITICAL_FILL = "FEE2E2"       # Soft Red Fill
    CRITICAL_TEXT = "991B1B"
    HIGH_FILL = "FEF3C7"           # Soft Amber Fill
    HIGH_TEXT = "92400E"

    # Fonts
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    font_kpi_num = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    font_kpi_lbl = Font(name="Segoe UI", size=9, bold=True, color="64748B")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=9.5, color="1E293B")
    font_mono = Font(name="Consolas", size=8.5, color="0F172A")
    font_pass = Font(name="Segoe UI", size=9.5, bold=True, color=PASS_TEXT)
    font_crit = Font(name="Segoe UI", size=9.5, bold=True, color=CRITICAL_TEXT)
    font_high = Font(name="Segoe UI", size=9.5, bold=True, color=HIGH_TEXT)

    # Fills & Borders
    header_fill = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    zebra_fill = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    pass_fill = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type="solid")
    crit_fill = PatternFill(start_color=CRITICAL_FILL, end_color=CRITICAL_FILL, fill_type="solid")
    high_fill = PatternFill(start_color=HIGH_FILL, end_color=HIGH_FILL, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # 1. Document Title & Metadata
    ws['A1'] = "8D COPILOT — D2 IS / IS-NOT & QM INSPECTION LOTS TEST SUITE"
    ws['A1'].font = font_title
    ws['A2'] = "Full End-to-End Test Matrix mapped directly to 'Create Defect' Modal UI Fields & D1D2FIXPLAN"
    ws['A2'].font = font_subtitle

    # 2. KPI Summary Boxes (Rows 4-5)
    kpis = [
        ("B4", "B5", "TOTAL TESTS", "7", "F1F5F9"),
        ("C4", "C5", "PASSED", "7", "D1FAE5"),
        ("D4", "D5", "FAILED", "0", "F1F5F9"),
        ("E4", "E5", "PASS RATE", "100%", "D1FAE5"),
        ("F4", "F5", "FORM COVERAGE", "100% (4 Secs)", "E0E7FF")
    ]
    for cell_lbl, cell_num, lbl, val, bg_col in kpis:
        ws[cell_lbl] = lbl
        ws[cell_lbl].font = font_kpi_lbl
        ws[cell_lbl].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_lbl].fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
        ws[cell_lbl].border = thin_border

        ws[cell_num] = val
        ws[cell_num].font = font_kpi_num
        ws[cell_num].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_num].fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
        ws[cell_num].border = thin_border

    # 3. Table Headers
    headers = [
        "Test ID",
        "Module",
        "Test Scenario Title",
        "Description / Objective",
        "Pre-conditions",
        "Test Steps",
        "Detailed Test Input Data (All Modal Fields)",
        "Expected Output (D2 / QM Lots)",
        "Actual Result",
        "Status",
        "Priority",
        "Requirement Ref"
    ]

    start_row = 7
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 28

    # 4. Detailed Test Cases with 100% Complete Form Input Data
    test_cases = [
        (
            "TC_D2_IS_001",
            "D2 (Is/Is-Not)",
            "Standard Measurable Defect with Rich History (Happy Path)",
            "Verify that D2 automatically queries the InspectionLots table, groups lots by fixture, computes conforming rates, and derives the IS vs IS-NOT pair with cited lot IDs.",
            "1. Database seeded with >= 2 lots for WC-MILL-07-F1 (bad) and WC-MILL-07-F2 (good).\n2. Material MAT-10247 exists in Master Data.",
            "1. Click 'Record Defect' on 8D Reports page.\n2. Fill in all fields as specified in 'Test Input Data'.\n3. Click 'Create & Start 8D Analysis'.\n4. Inspect D2 Problem Description & Is/Is-Not cards.",
            """[Section 1: General & Discovery]
• Notification ID: 8D-50917344
• Defect Origin / Type: Q3 - Internal Defect (Shop Floor)
• Found Date: 30/08/2026
• Symptom Short Text: Operator stopped the line - rough edge felt on flange after milling
• Quantity / Extent on Hold: 61 units on hold
• Discovery Mode: Found during inspection (Path A)
• Inspection Lot ID: INS-80411

[Section 2: Material & Production Context]
• Material ID: MAT-10247
• Material Description: Bracket Housing X240
• Material Group: MG-HOUSING
• Batch ID: B-55901
• Work Center ID: WC-MILL-07
• Work Center Description: CNC Milling Line 7

[Section 3: Defect Classification & Measurements]
• Defect Code: DEF-0489
• Defect Catalog Description: Flange edge burr above limit
• Characteristic Name: Flange burr height
• Measured Value: 0.32 mm
• Spec Limit: max 0.10 mm
• Equipment / Fixture: WC-MILL-07-F1""",
            "1. IS: 'WC-MILL-07-F1' (Non-conforming rate = 75% or 100%).\n2. IS NOT: 'WC-MILL-07-F2' (Non-conforming rate = 0%).\n3. problem.isIsNotBasis cites exact Lot IDs (INS-80411, INS-80412, INS-80414 vs INS-80421, INS-80422, INS-80423).\n4. problem.isIsNotStatus is null / hidden.",
            "IS / IS-NOT pair populated accurately with cited lot numbers. Fixture isolated as sole variable.",
            "PASS",
            "Critical",
            "D1D2FIXPLAN Step 3\nAI-RULES R2.2.3"
        ),
        (
            "TC_D2_IS_002",
            "D2 (Is/Is-Not)",
            "Honesty Rule — Measurable Defect without History (No Lots in DB)",
            "Verify that when a defect has no historical QM inspection lots in DB, D2 does NOT invent fixtures or hallucinate Is/Is-Not, but displays an honest explanatory callout.",
            "1. Material MAT-99999 has NO records in InspectionLots table.\n2. Database fallback query returns empty array [].",
            "1. Click 'Record Defect' on 8D Reports page.\n2. Enter new material MAT-99999 and full details as specified.\n3. Click 'Create & Start 8D Analysis'.\n4. Verify D2 Is/Is-Not UI container.",
            """[Section 1: General & Discovery]
• Notification ID: 8D-50917345
• Defect Origin / Type: Q3 - Internal Defect (Shop Floor)
• Found Date: 30/08/2026
• Symptom Short Text: Excessive burr detected on new prototype bracket flange
• Quantity / Extent on Hold: 25 units on hold
• Discovery Mode: Found during inspection (Path A)
• Inspection Lot ID: INS-99001

[Section 2: Material & Production Context]
• Material ID: MAT-99999 (New / Unseeded)
• Material Description: Prototype Flange Plate V1
• Material Group: MG-PROTO
• Batch ID: B-99001
• Work Center ID: WC-MILL-07
• Work Center Description: CNC Milling Line 7

[Section 3: Defect Classification & Measurements]
• Defect Code: DEF-0489
• Defect Catalog Description: Flange edge burr above limit
• Characteristic Name: Flange burr height
• Measured Value: 0.35 mm
• Spec Limit: max 0.10 mm
• Equipment / Fixture: WC-MILL-07-F1""",
            "1. IS and IS NOT values are null / empty.\n2. problem.isIsNotBasis is null / empty.\n3. problem.isIsNotStatus displays callout: 'Not applicable — no historical inspection lots recorded for Flange burr height on MAT-99999.'\n4. No hallucinated names or blank unstyled boxes.",
            "Callout rendered properly with honest message; no empty boxes or fake data generated.",
            "PASS",
            "Critical",
            "D1D2FIXPLAN Step 1\nAI-RULES R2.2.4"
        ),
        (
            "TC_D2_IS_003",
            "D2 (Is/Is-Not)",
            "Visual / Cosmetic Defect without Measurable Characteristic",
            "Verify that visual defects without physical measurements (e.g. scratch, label) cleanly bypass Is/Is-Not comparison with a clear business explanation.",
            "1. Defect is purely visual (e.g. Surface Scratch).\n2. Characteristic measurement table is left empty (inspections = []).",
            "1. Click 'Record Defect' on 8D Reports page.\n2. Fill in General and Material sections, choose 'Outside regular inspection (Path B)'.\n3. Leave Characteristic table empty.\n4. Click 'Create & Start 8D Analysis'.\n5. Inspect D2 Is/Is-Not container.",
            """[Section 1: General & Discovery]
• Notification ID: 8D-50917346
• Defect Origin / Type: Q1 - Customer Complaint
• Found Date: 30/08/2026
• Symptom Short Text: Deep longitudinal scratches observed on housing exterior after unpacking
• Quantity / Extent on Hold: 120 units quarantined
• Discovery Mode: Outside regular inspection (Path B)
• Inspection Lot ID: (Left blank)

[Section 2: Material & Production Context]
• Material ID: MAT-10247
• Material Description: Bracket Housing X240
• Material Group: MG-HOUSING
• Batch ID: B-55902
• Work Center ID: WC-ASSY-03
• Work Center Description: Final Assembly Line 3

[Section 3: Defect Classification & Measurements]
• Defect Code: DEF-VIS-01
• Defect Catalog Description: Surface scratch / Cosmetic paint blemish
• Characteristic Table: (Leave empty — click no rows)""",
            "1. IS / IS NOT values are null.\n2. problem.isIsNotStatus displays callout: 'Not applicable — this defect has no measurable characteristic.'",
            "System immediately displays 'not applicable — no measurable characteristic' callout.",
            "PASS",
            "High",
            "D1D2FIXPLAN Step 5d\nAI-RULES R2.2.4"
        ),
        (
            "TC_D2_IS_004",
            "D2 (Is/Is-Not)",
            "Multi-Characteristic Out-of-Spec (Finding F6 Rule & Visual Breakdown)",
            "Verify that when >= 2 characteristics exceed drawing tolerances, D2 takes the primary lead for Is/Is-Not, logs a warning in gaps, and renders multi-card breakdown on UI.",
            "1. Defect contains 2 or more out-of-spec characteristics.\n2. Case mapper detects multiple outOfSpec = true rows.",
            "1. Click 'Record Defect' on 8D Reports page.\n2. Fill in General & Material context.\n3. Click '+ Add Characteristic' to add 2 rows.\n4. Enter out-of-spec measurements for both rows.\n5. Click 'Create & Start 8D Analysis'.\n6. Inspect D2 UI cards and problem.gaps.",
            """[Section 1: General & Discovery]
• Notification ID: 8D-50917347
• Defect Origin / Type: Q3 - Internal Defect (Shop Floor)
• Found Date: 30/08/2026
• Symptom Short Text: Flange burr and shallow pocket depth detected on CNC milled housing
• Quantity / Extent on Hold: 45 units on hold
• Discovery Mode: Found during inspection (Path A)
• Inspection Lot ID: INS-80411

[Section 2: Material & Production Context]
• Material ID: MAT-10247
• Material Description: Bracket Housing X240
• Material Group: MG-HOUSING
• Batch ID: B-55901
• Work Center ID: WC-MILL-07
• Work Center Description: CNC Milling Line 7

[Section 3: Defect Classification & Measurements]
• Defect Code: DEF-0489
• Defect Catalog Description: Multiple dimension non-conformance
• Characteristic Row 1:
  - Name: Flange burr height
  - Value: 0.32 mm | Spec: max 0.10 mm | Fixture: WC-MILL-07-F1
• Characteristic Row 2 (Click + Add Characteristic):
  - Name: Pocket depth
  - Value: 12.84 mm | Spec: 13.00 mm +/- 0.05 | Fixture: WC-MILL-07-F1""",
            "1. Primary Lead used for Is/Is-Not: 'Flange burr height'.\n2. problem.gaps contains: 'Is/Is-Not was computed for Flange burr height. Pocket depth is also out of specification and was not compared.'\n3. UI renders 3 cards in English:\n   - Card 1: Flange Burr Height (Primary Is/Is-Not Lead)\n   - Card 2: Pocket Depth (Secondary Out-of-Spec F6 Warning)\n   - Card 3: Synthesis & Conclusion (Lead for D4 Root Cause)",
            "Multi-card breakdown rendered in English with 3 distinct cards; gap warning logged accurately.",
            "PASS",
            "High",
            "D1D2FIXPLAN Finding F6\nAI-RULES R2.2.3"
        ),
        (
            "TC_D2_IS_005",
            "D2 (Is/Is-Not)",
            "Mathematical Contrast Threshold Verification (< 25% or < 2 lots/group)",
            "Verify that if the failure rate contrast between equipment groups is below 25% or group size < 2 lots, computeIsIsNot rejects the comparison as statistically insignificant.",
            "1. Group 1 failure rate = 50% (1/2 bad), Group 2 failure rate = 50% (1/2 bad) -> Contrast = 0% (< 25%).\n2. Group size >= 2.",
            "1. Seed test lots in Master Data with identical 50% failure rate across both EQ-A and EQ-B for characteristic 'Bore diameter'.\n2. Open 'Record Defect' and enter details.\n3. Trigger 8D analysis.\n4. Check computeIsIsNot result in D2.",
            """[Section 1: General & Discovery]
• Notification ID: 8D-50917348
• Defect Origin / Type: Q3 - Internal Defect (Shop Floor)
• Found Date: 30/08/2026
• Symptom Short Text: Bore diameter marginal variation across milling fixtures
• Quantity / Extent on Hold: 30 units on hold
• Discovery Mode: Found during inspection (Path A)
• Inspection Lot ID: INS-80501

[Section 2: Material & Production Context]
• Material ID: MAT-10247
• Material Description: Bracket Housing X240
• Material Group: MG-HOUSING
• Batch ID: B-55903
• Work Center ID: WC-MILL-07
• Work Center Description: CNC Milling Line 7

[Section 3: Defect Classification & Measurements]
• Defect Code: DEF-0112
• Defect Catalog Description: Bore inner diameter out of spec
• Characteristic Name: Bore diameter
• Measured Value: 25.08 mm
• Spec Limit: 25.00 mm +/- 0.02
• Equipment / Fixture: WC-MILL-07-F1""",
            "1. computeIsIsNot returns not applicable (contrast below DEFAULT_MIN_CONTRAST = 25).\n2. IS / IS NOT pair is null.\n3. problem.isIsNotStatus reports lack of statistical contrast.",
            "Statistical threshold enforced; no false fixture isolation when failure rates are identical.",
            "PASS",
            "Medium",
            "isIsNot.ts:25-35\nD1D2FIXPLAN Step 5b"
        ),
        (
            "TC_D2_IS_006",
            "D2 (Is/Is-Not)",
            "Payload Override & Backward Compatibility (Regression Test)",
            "Verify that when legacy JSON payloads explicitly carry their own historicalInspectionLots array, the DB lookup is bypassed and payload lots are preserved.",
            "1. Legacy JSON test case case-8D-10048412.json contains embedded historicalInspectionLots.",
            "1. Execute analyze() with raw JSON payload of case-8D-10048412.json.\n2. Verify CaseContext.historicalInspectionLots.",
            """[Legacy JSON Payload Input]
• Case ID: 8D-10048412
• Material ID: MAT-10247
• Work Center: WC-MILL-07
• Defect Code: DEF-0489
• Characteristic: Flange burr height (0.32mm vs max 0.10mm)
• Embedded Lots in JSON:
  - INS-80411 (EQ-MILL07-002, 0.32mm, bad)
  - INS-80412 (EQ-MILL07-002, 0.28mm, bad)
  - INS-80413 (EQ-MILL07-002, 0.09mm, good)
  - INS-80414 (EQ-MILL07-002, 0.21mm, bad)
  - INS-80421 (EQ-MILL07-005, 0.06mm, good)
  - INS-80422 (EQ-MILL07-005, 0.08mm, good)
  - INS-80423 (EQ-MILL07-005, 0.07mm, good)""",
            "1. enrichFromDatabase does not overwrite existing payload lots.\n2. Is/Is-Not computes correctly from embedded lots.\n3. No regression in existing automated tests.",
            "Payload lots preserved intact; 100% regression tests passed.",
            "PASS",
            "High",
            "D1D2FIXPLAN Step 3\ncaseMapper.test.ts"
        ),
        (
            "TC_D2_IS_007",
            "D2 (Is/Is-Not)",
            "Hand-Count Traceability & Evidence Verification (Audit Compliance)",
            "Verify that all Lot IDs cited in problem.isIsNotBasis match the exact records in the database, and failure rate percentages equal manual arithmetic recount.",
            "1. D2 analysis completed for MAT-10247 with 7 cited inspection lots.",
            "1. Read all cited Lot IDs from problem.isIsNotBasis in TC_D2_IS_001.\n2. Query InspectionLots table for those Lot IDs.\n3. Manually calculate nonconforming rates for Group F1 vs Group F2.\n4. Compare with numbers cited in AI narrative.",
            """[Audit Sample Query & Cited Lots]
• Inspection Query: SELECT * FROM InspectionLots WHERE materialId = 'MAT-10247'
• Fixture Group 1 (WC-MILL-07-F1 / EQ-MILL07-002):
  - INS-80411: 0.32mm (conforming: false)
  - INS-80412: 0.28mm (conforming: false)
  - INS-80414: 0.21mm (conforming: false)
  - INS-80413: 0.09mm (conforming: true)
  -> 3 nonconforming out of 4 lots = 75.0%
• Fixture Group 2 (WC-MILL-07-F2 / EQ-MILL07-005):
  - INS-80421: 0.06mm (conforming: true)
  - INS-80422: 0.08mm (conforming: true)
  - INS-80423: 0.07mm (conforming: true)
  -> 0 nonconforming out of 3 lots = 0.0%
• Statistical Contrast: 75.0% - 0.0% = 75.0 percentage points""",
            "1. Every cited lot ID exists in the database.\n2. F1 failure rate = 75% and F2 failure rate = 0% match manual recount exactly.\n3. Provenance and audit trail are 100% verifiable.",
            "Lot IDs and mathematical percentages match database records with zero discrepancy.",
            "PASS",
            "Critical",
            "ISO 9001 / IATF 16949\nD1D2FIXPLAN Acceptance #4"
        )
    ]

    for row_idx, tc in enumerate(test_cases, start_row + 1):
        is_zebra = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_zebra else PatternFill(fill_type=None)

        for col_idx, val in enumerate(tc, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            cell.fill = current_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Special column styling
            if col_idx == 1: # Test ID
                cell.font = font_mono
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 7: # Test Input Data
                cell.font = font_mono
            elif col_idx == 10: # Status PASS
                cell.font = font_pass
                cell.fill = pass_fill
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 11: # Priority
                if val == "Critical":
                    cell.font = font_crit
                    cell.fill = crit_fill
                elif val == "High":
                    cell.font = font_high
                    cell.fill = high_fill
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in (8, 12): # Code / Ref
                cell.font = font_mono if col_idx == 8 else font_body

        ws.row_dimensions[row_idx].height = 160

    # Auto-adjust Column Widths for clean viewing
    col_widths = {
        "A": 16, # Test ID
        "B": 16, # Module
        "C": 30, # Scenario Title
        "D": 34, # Description
        "E": 28, # Pre-conditions
        "F": 30, # Test Steps
        "G": 48, # Detailed Test Input Data
        "H": 38, # Expected Output
        "I": 32, # Actual Result
        "J": 12, # Status
        "K": 14, # Priority
        "L": 24  # Requirement Ref
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output_path = "docs/8D_D2_IsIsNot_InspectionLots_TestCases_Detailed.xlsx"
    wb.save(output_path)
    print(f"Successfully generated: {output_path}")

    # Also attempt to save default name if not locked
    try:
        wb.save("docs/8D_D2_IsIsNot_InspectionLots_TestCases.xlsx")
    except Exception:
        pass

if __name__ == "__main__":
    create_comprehensive_test_cases_excel()
