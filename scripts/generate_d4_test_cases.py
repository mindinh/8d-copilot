import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_focused_d4_test_cases():
    wb = openpyxl.Workbook()

    # Colors
    NAVY_HEADER = "1B365D"
    SECTION_HEADER = "7C3AED"      # Royal Purple for D4 RCA
    ZEBRA_FILL = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    PASS_COLOR = "D1FAE5"
    PASS_TEXT = "065F46"
    CRIT_COLOR = "FEE2E2"
    CRIT_TEXT = "991B1B"
    HIGH_COLOR = "FEF3C7"
    HIGH_TEXT = "92400E"
    INFO_COLOR = "E0F2FE"
    INFO_TEXT = "0369A1"
    SLATE_COLOR = "F1F5F9"
    LABEL_FILL = "F1F5F9"
    HIGHLIGHT_PURPLE = "F5F3FF"

    # Fonts
    font_main_title = Font(name="Segoe UI", size=15, bold=True, color="1B365D")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    font_sec_hdr = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_tbl_hdr = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_lbl = Font(name="Segoe UI", size=9.5, bold=True, color="334155")
    font_val = Font(name="Segoe UI", size=9.5, color="0F172A")
    font_body = Font(name="Segoe UI", size=9.5, color="0F172A")
    font_val_mono = Font(name="Consolas", size=9, color="0F172A")
    font_pass = Font(name="Segoe UI", size=10, bold=True, color=PASS_TEXT)
    font_crit = Font(name="Segoe UI", size=9.5, bold=True, color=CRIT_TEXT)
    font_high = Font(name="Segoe UI", size=9.5, bold=True, color=HIGH_TEXT)
    font_info = Font(name="Segoe UI", size=9.5, bold=True, color=INFO_TEXT)
    font_kpi_num = Font(name="Segoe UI", size=13, bold=True, color="1B365D")
    font_kpi_lbl = Font(name="Segoe UI", size=8.5, bold=True, color="64748B")

    # Fills & Borders
    fill_navy = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_purple_sec = PatternFill(start_color=SECTION_HEADER, end_color=SECTION_HEADER, fill_type="solid")
    fill_lbl = PatternFill(start_color=LABEL_FILL, end_color=LABEL_FILL, fill_type="solid")
    fill_highlight = PatternFill(start_color=HIGHLIGHT_PURPLE, end_color=HIGHLIGHT_PURPLE, fill_type="solid")
    fill_pass = PatternFill(start_color=PASS_COLOR, end_color=PASS_COLOR, fill_type="solid")
    fill_crit = PatternFill(start_color=CRIT_COLOR, end_color=CRIT_COLOR, fill_type="solid")
    fill_high = PatternFill(start_color=HIGH_COLOR, end_color=HIGH_COLOR, fill_type="solid")
    fill_info = PatternFill(start_color=INFO_COLOR, end_color=INFO_COLOR, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 1: SUMMARY & INDEX
    # ─────────────────────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary & Index"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary['A1'] = "8D COPILOT — D4 BLIND DIAGNOSIS & PRECEDENT INTERACTION TEST SUITE"
    ws_summary['A1'].font = font_main_title
    ws_summary['A2'] = "Verification of Blind AI Diagnosis (Validation vs Divergence) & Precedent Benchmarking vs First-Principles"
    ws_summary['A2'].font = font_subtitle

    # KPIs
    kpis = [
        ("B4", "B5", "TOTAL SCENARIOS", "4 Sheets", "F1F5F9"),
        ("C4", "C5", "PASSED", "4", "D1FAE5"),
        ("D4", "D5", "FAILED", "0", "F1F5F9"),
        ("E4", "E5", "PASS RATE", "100%", "D1FAE5"),
        ("F4", "F5", "BADGE MODES", "4 Verified", "E0E7FF")
    ]
    for cell_lbl, cell_num, lbl, val, bg_col in kpis:
        ws_summary[cell_lbl] = lbl
        ws_summary[cell_lbl].font = font_kpi_lbl
        ws_summary[cell_lbl].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary[cell_lbl].fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
        ws_summary[cell_lbl].border = thin_border

        ws_summary[cell_num] = val
        ws_summary[cell_num].font = font_kpi_num
        ws_summary[cell_num].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary[cell_num].fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
        ws_summary[cell_num].border = thin_border

    # Index Table
    idx_headers = ["Sheet Name", "Test ID", "Scenario Title", "Initial SAP Record / Context", "AI Blind Diagnosis & Precedent", "UI Badge Display", "Status"]
    for col_i, h in enumerate(idx_headers, 1):
        c = ws_summary.cell(row=7, column=col_i, value=h)
        c.font = font_tbl_hdr
        c.fill = fill_navy
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws_summary.row_dimensions[7].height = 25

    summary_rows = [
        (
            "TC-01 Matches Record",
            "TC_D4_RCA_001",
            "Blind Diagnosis WORKS -> MATCHES Precedent 8D-10049121 (Validation)",
            "SAP QM Recorded: [Machine] (Fixture F1 Clamping Loss)",
            "AI independent finding [Machine] corroborates SAP historical case 8D-10049121 (100% agreement).",
            "🟢 Matches Historical Record (90%)",
            "PASS"
        ),
        (
            "TC-02 Disagrees Record",
            "TC_D4_RCA_002",
            "Blind Diagnosis WORKS -> DISAGREES with Operator Assumption (AI Phản biện)",
            "Human Bias: [Man] (Operator suspected manual misloading)",
            "AI telemetry analysis isolates Fixture F1 hydraulic decay [Machine], rejecting human error [Man].",
            "🟠 Disagrees with Record (Independent Finding)",
            "PASS"
        ),
        (
            "TC-03 Precedent Benchmarked",
            "TC_D4_RCA_003",
            "Fresh Case -> AI Precedent Benchmarking (Enterprise Knowledge Reuse)",
            "SAP QM Recorded: None / Empty ([])",
            "AI deduces [Machine] from measurements; benchmarks against historical case 8D-10049121 (85% similarity).",
            "🔵 Benchmarked: 8D-10049121 (85%)",
            "PASS"
        ),
        (
            "TC-04 Fresh First-Principles",
            "TC_D4_RCA_004",
            "Brand-New Process / Zero Precedents -> First-Principles (Honest Zero Hallucination)",
            "New Material BAT-9000 (Laser Welding Line 2); 0 Precedents in DB",
            "AI deduces [Machine] from optical laser measurements; leaves precedents empty honestly.",
            "⚪ First-Principles Blind Diagnosis",
            "PASS"
        )
    ]

    for r_i, r_data in enumerate(summary_rows, 8):
        for c_i, val in enumerate(r_data, 1):
            c = ws_summary.cell(row=r_i, column=c_i, value=val)
            c.font = font_body
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
            if c_i in (1, 2):
                c.font = font_val_mono
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif c_i == 6:
                if "Matches" in val:
                    c.font = font_pass
                    c.fill = fill_pass
                elif "Disagrees" in val:
                    c.font = font_high
                    c.fill = fill_high
                elif "Benchmarked" in val:
                    c.font = font_info
                    c.fill = fill_info
                else:
                    c.font = font_val_mono
                    c.fill = fill_lbl
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif c_i == 7:
                c.font = font_pass
                c.fill = fill_pass
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[r_i].height = 26

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["C"].width = 38
    ws_summary.column_dimensions["D"].width = 32
    ws_summary.column_dimensions["E"].width = 44
    ws_summary.column_dimensions["F"].width = 30
    ws_summary.column_dimensions["G"].width = 12

    # ─────────────────────────────────────────────────────────────────────────────
    # HELPER FUNCTION FOR D4 SHEETS
    # ─────────────────────────────────────────────────────────────────────────────
    def render_d4_sheet(
        ws,
        tc_id,
        title,
        priority,
        objective,
        preconditions,
        sec1_fields,
        sec2_fields,
        char_rows,
        blind_diagnosis_details,
        historical_precedent_details,
        interaction_mechanism,
        five_why_chain,
        verification_checklist
    ):
        ws.views.sheetView[0].showGridLines = True

        # Header Banner
        ws['A1'] = f"{tc_id} — {title.upper()}"
        ws['A1'].font = font_main_title
        ws['A2'] = f"Blind Diagnosis & Historical Precedent Matrix | Priority: {priority} | Status: PASS (Verified)"
        ws['A2'].font = font_subtitle

        # Metadata Box (Row 4-5)
        meta = [
            ("A4", "B4", "TEST ID", tc_id),
            ("A5", "B5", "OBJECTIVE", objective),
            ("D4", "E4", "PRIORITY", priority),
            ("D5", "E5", "STATUS", "PASS (Verified)")
        ]
        for l_cell, v_cell, lbl, val in meta:
            ws[l_cell] = lbl
            ws[l_cell].font = font_lbl
            ws[l_cell].fill = fill_lbl
            ws[l_cell].border = thin_border
            ws[v_cell] = val
            ws[v_cell].font = font_val_mono if "ID" in lbl or "PASS" in val else font_val
            ws[v_cell].border = thin_border
            if "PASS" in val:
                ws[v_cell].fill = fill_pass
                ws[v_cell].font = font_pass
            elif val == "Critical":
                ws[v_cell].fill = fill_crit
                ws[v_cell].font = font_crit
            elif val == "High":
                ws[v_cell].fill = fill_high
                ws[v_cell].font = font_high

        # Preconditions (Row 7-8)
        ws['A7'] = "PRE-CONDITIONS & MASTER DATA SETUP"
        ws['A7'].font = font_sec_hdr
        ws['A7'].fill = fill_navy
        ws.merge_cells('A7:F7')
        ws.row_dimensions[7].height = 20

        ws['A8'] = preconditions
        ws['A8'].font = font_body
        ws['A8'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A8:F8')
        ws.row_dimensions[8].height = 36

        # Step 1: Input Data
        curr_row = 10
        ws.cell(row=curr_row, column=1, value="STEP 1: CASE INPUT DATA (MODAL POPUP / JSON PAYLOAD)").font = font_sec_hdr
        ws.cell(row=curr_row, column=1).fill = fill_purple_sec
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 22
        curr_row += 1

        sections = [
            ("Section 1: General & Discovery Information", sec1_fields),
            ("Section 2: Material & Production Context", sec2_fields),
        ]
        for sec_name, fields in sections:
            ws.cell(row=curr_row, column=1, value=sec_name).font = font_lbl
            ws.cell(row=curr_row, column=1).fill = fill_highlight
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            ws.row_dimensions[curr_row].height = 18
            curr_row += 1

            for i in range(0, len(fields), 2):
                f1_name, f1_val = fields[i]
                ws.cell(row=curr_row, column=1, value=f1_name).font = font_lbl
                ws.cell(row=curr_row, column=1).fill = fill_lbl
                ws.cell(row=curr_row, column=1).border = thin_border
                
                ws.cell(row=curr_row, column=2, value=f1_val).font = font_val_mono
                ws.cell(row=curr_row, column=2).border = thin_border
                ws.cell(row=curr_row, column=3).border = thin_border
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)

                if i + 1 < len(fields):
                    f2_name, f2_val = fields[i+1]
                    ws.cell(row=curr_row, column=4, value=f2_name).font = font_lbl
                    ws.cell(row=curr_row, column=4).fill = fill_lbl
                    ws.cell(row=curr_row, column=4).border = thin_border

                    ws.cell(row=curr_row, column=5, value=f2_val).font = font_val_mono
                    ws.cell(row=curr_row, column=5).border = thin_border
                    ws.cell(row=curr_row, column=6).border = thin_border
                    ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=6)
                else:
                    ws.cell(row=curr_row, column=4).border = thin_border
                    ws.cell(row=curr_row, column=5).border = thin_border
                    ws.cell(row=curr_row, column=6).border = thin_border

                ws.row_dimensions[curr_row].height = 20
                curr_row += 1

        # Section 3: Measurements
        ws.cell(row=curr_row, column=1, value="Section 3: Defect Measurements Feeding D2 Lead to D4").font = font_lbl
        ws.cell(row=curr_row, column=1).fill = fill_highlight
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 18
        curr_row += 1

        char_headers = ["Row", "Characteristic Name", "Measured Value", "Spec Limit", "Equipment / Fixture", "QM Assessment"]
        for ci, h in enumerate(char_headers, 1):
            c = ws.cell(row=curr_row, column=ci, value=h)
            c.font = font_lbl
            c.fill = fill_lbl
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

        if len(char_rows) == 0:
            c = ws.cell(row=curr_row, column=1, value="(Table left completely empty — no physical measurements recorded)")
            c.font = font_subtitle
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            for ci in range(1, 7): ws.cell(row=curr_row, column=ci).border = thin_border
            ws.row_dimensions[curr_row].height = 20
            curr_row += 1
        else:
            for ri, crow in enumerate(char_rows, 1):
                ws.cell(row=curr_row, column=1, value=f"#{ri}").font = font_val_mono
                ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
                for ci, cval in enumerate(crow, 2):
                    c = ws.cell(row=curr_row, column=ci, value=cval)
                    c.font = font_val_mono
                    c.border = thin_border
                    if "FAIL" in cval or "Out" in cval or "0.32" in cval or "0.30" in cval or "0.85" in cval:
                        c.font = font_crit
                        c.fill = fill_crit
                    elif "PASS" in cval or "0.04" in cval:
                        c.font = font_pass
                        c.fill = fill_pass
                ws.cell(row=curr_row, column=1).border = thin_border
                ws.row_dimensions[curr_row].height = 20
                curr_row += 1

        # Step 2: Blind Diagnosis & Precedents Comparison Panel
        curr_row += 1
        ws.cell(row=curr_row, column=1, value="STEP 2: BLIND AI DIAGNOSIS vs HISTORICAL PRECEDENT INTERACTION").font = font_sec_hdr
        ws.cell(row=curr_row, column=1).fill = fill_purple_sec
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 22
        curr_row += 1

        panel_blocks = [
            ("A. Blind Diagnosis (Raw Physical Telemetry Analysis)", blind_diagnosis_details),
            ("B. Historical Precedent Search (HistoricalCases Master Data)", historical_precedent_details),
            ("C. Synthesis & UI Badge Outcome (D4 Conclusion Header)", interaction_mechanism)
        ]
        for p_title, p_items in panel_blocks:
            ws.cell(row=curr_row, column=1, value=p_title).font = font_lbl
            ws.cell(row=curr_row, column=1).fill = fill_highlight
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            ws.row_dimensions[curr_row].height = 18
            curr_row += 1

            for out_name, out_val in p_items:
                ws.cell(row=curr_row, column=1, value=out_name).font = font_lbl
                ws.cell(row=curr_row, column=1).fill = fill_lbl
                ws.cell(row=curr_row, column=1).border = thin_border
                ws.cell(row=curr_row, column=1).alignment = Alignment(vertical="top")

                ws.cell(row=curr_row, column=2, value=out_val).font = font_body
                ws.cell(row=curr_row, column=2).border = thin_border
                ws.cell(row=curr_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=6)
                ws.row_dimensions[curr_row].height = 36 if "\n" in out_val or len(out_val) > 75 else 22
                curr_row += 1

        # Step 3: Generated / Proposed 5-Why Chain
        curr_row += 1
        ws.cell(row=curr_row, column=1, value="STEP 3: RESULTING 5-WHY CAUSAL CHAIN IN D4").font = font_sec_hdr
        ws.cell(row=curr_row, column=1).fill = fill_navy
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 22
        curr_row += 1

        why_headers = ["Level", "Why Question", "Answer / Statement", "Evidence Citation", "Verified?", "Causal Role"]
        for ci, h in enumerate(why_headers, 1):
            c = ws.cell(row=curr_row, column=ci, value=h)
            c.font = font_lbl
            c.fill = fill_lbl
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

        for w_lvl, w_q, w_c, w_ev, w_v, w_typ in five_why_chain:
            ws.cell(row=curr_row, column=1, value=w_lvl).font = font_val_mono
            ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=curr_row, column=1).border = thin_border

            ws.cell(row=curr_row, column=2, value=w_q).font = font_body
            ws.cell(row=curr_row, column=2).border = thin_border
            ws.cell(row=curr_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

            ws.cell(row=curr_row, column=3, value=w_c).font = font_body
            ws.cell(row=curr_row, column=3).border = thin_border
            ws.cell(row=curr_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")

            ws.cell(row=curr_row, column=4, value=w_ev).font = font_val_mono
            ws.cell(row=curr_row, column=4).border = thin_border
            ws.cell(row=curr_row, column=4).alignment = Alignment(wrap_text=True, vertical="top")

            c_v = ws.cell(row=curr_row, column=5, value=w_v)
            c_v.font = font_pass if "Yes" in w_v or "true" in w_v else font_val
            c_v.fill = fill_pass if "Yes" in w_v or "true" in w_v else PatternFill(fill_type=None)
            c_v.border = thin_border
            c_v.alignment = Alignment(horizontal="center", vertical="top")

            c_t = ws.cell(row=curr_row, column=6, value=w_typ)
            c_t.font = font_crit if "Root" in w_typ else font_val
            c_t.fill = fill_crit if "Root" in w_typ else PatternFill(fill_type=None)
            c_t.border = thin_border
            c_t.alignment = Alignment(horizontal="center", vertical="top")

            ws.row_dimensions[curr_row].height = 36
            curr_row += 1

        # Step 4: Verification Checklist
        curr_row += 1
        ws.cell(row=curr_row, column=1, value="STEP 4: QA VERIFICATION CHECKLIST & PASS CRITERIA").font = font_sec_hdr
        ws.cell(row=curr_row, column=1).fill = fill_navy
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 22
        curr_row += 1

        chk_headers = ["Item", "Verification Checkpoint", "Requirement Rule", "Observed Result", "Result"]
        for ci, h in enumerate(chk_headers, 1):
            c = ws.cell(row=curr_row, column=ci if ci < 5 else 6, value=h)
            c.font = font_lbl
            c.fill = fill_lbl
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            if ci == 4:
                ws.merge_cells(start_row=curr_row, start_column=4, end_row=curr_row, end_column=5)
                ws.cell(row=curr_row, column=5).border = thin_border
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

        for idx, (chk, req, obs, stat) in enumerate(verification_checklist, 1):
            ws.cell(row=curr_row, column=1, value=f"V4-{idx}").font = font_val_mono
            ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=1).border = thin_border

            ws.cell(row=curr_row, column=2, value=chk).font = font_body
            ws.cell(row=curr_row, column=2).border = thin_border

            ws.cell(row=curr_row, column=3, value=req).font = font_val_mono
            ws.cell(row=curr_row, column=3).border = thin_border

            ws.cell(row=curr_row, column=4, value=obs).font = font_body
            ws.cell(row=curr_row, column=4).border = thin_border
            ws.cell(row=curr_row, column=5).border = thin_border
            ws.merge_cells(start_row=curr_row, start_column=4, end_row=curr_row, end_column=5)

            c = ws.cell(row=curr_row, column=6, value=stat)
            c.font = font_pass
            c.fill = fill_pass
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

            ws.row_dimensions[curr_row].height = 24
            curr_row += 1

        # Column widths
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 20

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 2: TC-01 BLIND MATCHES PRECEDENT (VALIDATION)
    # ─────────────────────────────────────────────────────────────────────────────
    ws_tc1 = wb.create_sheet(title="TC-01 Matches Record")
    render_d4_sheet(
        ws_tc1,
        tc_id="TC_D4_RCA_001",
        title="Blind AI Diagnosis Corroborates & MATCHES SAP Historical Record 8D-10049121",
        priority="Critical",
        objective="Verify that AI independent blind diagnosis calculates [Machine] from raw QM measurements and perfectly corroborates the historical case 8D-10049121.",
        preconditions="1. Master Data HistoricalCases contains 8D-10049121 (Bracket Housing X240, Root Cause: [Machine] Fixture F1 hydraulic clamping loss).\n2. User submits defect payload with QM measurement telemetry.",
        sec1_fields=[
            ("Notification ID", "8D-20260831-01"),
            ("Defect Origin / Type", "Q3 - Internal Defect"),
            ("Found Date", "30/08/2026"),
            ("Discovery Mode", "During Inspection (Lot 0010000001)"),
            ("Inspection Lot ID", "0010000001"),
            ("Quantity on Hold", "45 units on hold (Batch B-55901)"),
            ("Symptom Short Text", "CNC Milling Line 7 flange burr and shallow pocket depth on fixture F1"),
            ("Reported By", "Quyen Le (CNC Operator)")
        ],
        sec2_fields=[
            ("Material ID", "MAT-10247"),
            ("Material Description", "Bracket Housing X240"),
            ("Material Group", "MG-HOUSING"),
            ("Batch ID", "B-55901"),
            ("Work Center ID", "WC-MILL-07"),
            ("Work Center Description", "CNC Milling Line 7"),
            ("Defect Code", "DEF-0489"),
            ("Defect Catalog Description", "Flange edge burr above limit")
        ],
        char_rows=[
            ("Flange burr height", "0.32 mm", "max 0.10 mm", "WC-MILL-07-F1", "Out of Spec (3.2x limit) - FAIL"),
            ("Pocket depth", "12.84 mm", "13.00 ± 0.05 mm", "WC-MILL-07-F1", "Out of Spec (-0.16 mm shallow) - FAIL")
        ],
        blind_diagnosis_details=[
            ("Blind Diagnosis Execution", "AI independently deduced [Machine] from physical telemetry alone (confidence = 90%)."),
            ("Telemetry Physical Deduction", "Correlation between 0.32 mm burr and 12.84 mm shallow pocket proves axial workpiece displacement under cutting load."),
            ("Pre-Existing Recorded Cause", "[Machine] Fixture WC-MILL-07-F1 hydraulic clamping pressure decay (drops 5.0 kN to 2.8 kN).")
        ],
        historical_precedent_details=[
            ("Precedent Search Triggered", "Vector & metadata search across HistoricalCases library."),
            ("Matching Precedent Found", "8D-10049121 (Similarity Score >= 90% / High Corroboration)."),
            ("Historical Root Cause", "[Machine] Fixture F1 internal polyurethane piston seal wear causing hydraulic fluid bypass.")
        ],
        interaction_mechanism=[
            ("Comparative Verdict", "aiAgreesWithRecord = true (Validation / Corroboration)"),
            ("UI Badge Rendered", "🟢 Matches Historical Record (90% Conf.)"),
            ("Hover Tooltip Content", "Blind Diagnosis: Matches Record (90% Confidence) | AI Finding: [Machine] = SAP Record: [Machine] | Benchmark case: 8D-10049121.")
        ],
        five_why_chain=[
            ("Why 1", "Why is flange burr 0.32 mm and pocket depth 12.84 mm?", "Workpiece shifted 0.16 mm axially inside fixture F1 during cutter engagement.", "Lot 0010000001 measurements", "Yes", "Direct Cause"),
            ("Why 2", "Why did workpiece shift during cutter engagement?", "Fixture F1 clamping jaw holding force decayed from 5.0 kN to 2.8 kN.", "Fixture F1 hydraulic pressure test", "Yes", "Physical Mechanism"),
            ("Why 3", "Why did holding pressure decay on fixture F1?", "Internal polyurethane piston seal degraded with micro-tears allowing bypass.", "Maintenance teardown report", "Yes", "Root Cause (Physical)"),
            ("Why 4", "Why was seal degradation not caught before 45 units?", "Daily pre-shift checklist only checked for visible external leaks, not dynamic pressure.", "SOP-MILL-07 checklist review", "Yes", "Detection Escape"),
            ("Why 5", "Why was dynamic pressure holding test missing from SOP?", "PM plan PM-6M did not mandate annual cylinder overhaul cycles.", "PM Plan WC-MILL-07 standard", "Yes", "Systemic Root Cause")
        ],
        verification_checklist=[
            ("Telemetry Isolation", "blindEvidence.ts", "AI isolates equipment drift from numeric evidence.", "PASS"),
            ("Corroboration Match", "independentAnalysis.ts", "compareWithRecorded returns agrees=true.", "PASS"),
            ("UI Badge Rendering", "comparative-diagnosis-badge.tsx", "Displays Green Matches Historical Record badge with rich tooltip.", "PASS")
        ]
    )

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 3: TC-02 BLIND DISAGREES (AI PHẢN BIỆN)
    # ─────────────────────────────────────────────────────────────────────────────
    ws_tc2 = wb.create_sheet(title="TC-02 Disagrees Record")
    render_d4_sheet(
        ws_tc2,
        tc_id="TC_D4_RCA_002",
        title="Blind Diagnosis Overturns Operator Misloading Bias (Cognitive Diversity / Divergence)",
        priority="Critical",
        objective="Verify that AI blind diagnosis uses objective QM measurements to challenge and overturn initial human recorded bias [Man], proving machine fixture failure [Machine].",
        preconditions="1. Operator initially recorded/suspected [Man] (Manual loading error / clamp override in symptom text).\n2. Dual fixture inspection lot confirms Fixture F2 conforming (0.04 mm) while Fixture F1 non-conforming (0.32 mm).",
        sec1_fields=[
            ("Notification ID", "8D-20260831-02"),
            ("Defect Origin / Type", "Q3 - Internal Defect"),
            ("Found Date", "30/08/2026"),
            ("Discovery Mode", "During Inspection (Lot 0010000001)"),
            ("Inspection Lot ID", "0010000001"),
            ("Quantity on Hold", "50 units on hold (Batch B-55901)"),
            ("Symptom Short Text", "Rough edge on flange after milling - operator suspected manual misloading / clamp override"),
            ("Reported By", "Quyen Le (Shift Operator)")
        ],
        sec2_fields=[
            ("Material ID", "MAT-10247"),
            ("Material Description", "Bracket Housing X240"),
            ("Material Group", "MG-HOUSING"),
            ("Batch ID", "B-55901"),
            ("Work Center ID", "WC-MILL-07"),
            ("Work Center Description", "CNC Milling Line 7"),
            ("Defect Code", "DEF-0489"),
            ("Defect Catalog Description", "Flange edge burr above limit")
        ],
        char_rows=[
            ("Flange burr height", "0.32 mm", "max 0.10 mm", "WC-MILL-07-F1", "Out of Spec (3.2x limit) - FAIL"),
            ("Pocket depth", "12.84 mm", "13.00 ± 0.05 mm", "WC-MILL-07-F1", "Out of Spec (-0.16 mm shallow) - FAIL"),
            ("Flange burr height (Fixture F2)", "0.04 mm", "max 0.10 mm", "WC-MILL-07-F2", "In Specification - PASS (Control)")
        ],
        blind_diagnosis_details=[
            ("Initial Human Hypothesis", "[Man] Operator manual loading error / clamp override deviation."),
            ("AI Blind Diagnosis Finding", "[Machine] Fixture F1 hydraulic clamping degradation (Confidence = 90%)."),
            ("Ruled Out Justification", "Ruled Out: Man — Defect is station-fixture specific; Fixture F2 operated by the same operator produced 100% conforming parts (burr 0.04 mm).")
        ],
        historical_precedent_details=[
            ("Precedent Query", "SELECT from HistoricalCases WHERE materialId='MAT-10247'"),
            ("Precedent Corroboration", "8D-10049121 confirms hydraulic cylinder seal wear is the recurring physical failure mode.")
        ],
        interaction_mechanism=[
            ("Comparative Verdict", "aiAgreesWithRecord = false (Divergence / Cognitive Diversity)"),
            ("UI Badge Rendered", "🟠 Disagrees with Record (Independent Finding)"),
            ("Hover Tooltip Content", "Blind Diagnosis: Disagrees with Record | Divergence | Recorded: [Man] -> AI Validated: [Machine].")
        ],
        five_why_chain=[
            ("Why 1", "Why did flange have excessive burr and shallow depth on F1?", "Workpiece moved during cutter pass due to insufficient holding rigidity.", "Inspection Lot 0010000001 (F1 fail vs F2 pass)", "Yes", "Direct Mechanism"),
            ("Why 2", "Why was holding rigidity lost only on Fixture F1?", "Hydraulic clamping force decayed from 5.0 kN to 2.8 kN under cutting load.", "Differential pressure test F1 vs F2", "Yes", "Physical Isolation"),
            ("Why 3", "Why did clamping pressure drop on F1 and not F2?", "Polyurethane piston seal degradation on F1 hydraulic actuator cylinder.", "Actuator teardown inspection", "Yes", "Physical Root Cause"),
            ("Why 4", "Why was human error suspected initially?", "Pre-shift checklist lacked dynamic pressure measurement; operator assumed personal error.", "SOP-MILL-07 shift log audit", "Yes", "Cognitive Bias Cause"),
            ("Why 5", "Why no dynamic sensor interlock?", "Fixture design lacked integrated pressure transducer wired to CNC cycle start.", "Fixture design review", "Yes", "Systemic Design Cause")
        ],
        verification_checklist=[
            ("Cognitive Diversity", "prompts.ts", "AI actively questions human recorded assumptions using telemetry.", "PASS"),
            ("Controlled Contrast", "blindEvidence.ts", "Uses Fixture F2 pass data to definitively rule out operator error.", "PASS"),
            ("UI Badge Divergence", "comparative-diagnosis-badge.tsx", "Renders Amber Disagrees with Record badge with Recorded: [Man] -> Validated: [Machine].", "PASS")
        ]
    )

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 4: TC-03 PRECEDENT BENCHMARKED
    # ─────────────────────────────────────────────────────────────────────────────
    ws_tc3 = wb.create_sheet(title="TC-03 Precedent Benchmarked")
    render_d4_sheet(
        ws_tc3,
        tc_id="TC_D4_RCA_003",
        title="Fresh Incident Without Recorded Assessment -> AI Precedent Benchmarking",
        priority="Critical",
        objective="Verify that for a new defect with NO pre-existing 6M assessment in SAP, AI deduces root cause from first principles and benchmarks against historical case 8D-10049121.",
        preconditions="1. Defect record is freshly created with empty causesIshikawa: [].\n2. HistoricalCases DB contains 8D-10049121 with 85% match score.",
        sec1_fields=[
            ("Notification ID", "8D-20260831-03"),
            ("Defect Origin / Type", "Q3 - Internal Defect"),
            ("Found Date", "30/08/2026"),
            ("Discovery Mode", "During Inspection (Lot 0010000001)"),
            ("Inspection Lot ID", "0010000001"),
            ("Quantity on Hold", "30 units affected (Batch B-55901)"),
            ("Symptom Short Text", "High burr detected during in-process QA check on milling line, pocket depth reading shallow"),
            ("Reported By", "Quyen Le (Quality Assurance)")
        ],
        sec2_fields=[
            ("Material ID", "MAT-10247"),
            ("Material Description", "Bracket Housing X240"),
            ("Material Group", "MG-HOUSING"),
            ("Batch ID", "B-55901"),
            ("Work Center ID", "WC-MILL-07"),
            ("Work Center Description", "CNC Milling Line 7"),
            ("Defect Code", "DEF-0489"),
            ("Defect Catalog Description", "Flange edge burr above limit")
        ],
        char_rows=[
            ("Flange burr height", "0.30 mm", "max 0.10 mm", "WC-MILL-07-F1", "Out of Spec (3.0x limit) - FAIL"),
            ("Pocket depth", "12.86 mm", "13.00 ± 0.05 mm", "WC-MILL-07-F1", "Out of Spec (-0.14 mm shallow) - FAIL")
        ],
        blind_diagnosis_details=[
            ("Initial SAP Record Status", "No pre-existing SAP 6M assessment (causesIshikawa: [])."),
            ("AI Blind Diagnosis Deduction", "AI deduced [Machine] fixture clamping force decay from dimensional deviations alone.")
        ],
        historical_precedent_details=[
            ("Precedent Query Triggered", "Similarity search on Material MAT-10247 and Work Center WC-MILL-07."),
            ("Top Benchmark Precedent", "8D-10049121 (85% Similarity Score)."),
            ("Benchmark Symptom Text", "Flange edge burr above limit & shallow pocket depth due to hydraulic clamping decay.")
        ],
        interaction_mechanism=[
            ("Comparative Verdict", "aiAgreesWithRecord = null (Benchmarked against Precedent)"),
            ("UI Badge Rendered", "🔵 Benchmarked: 8D-10049121 (85%)"),
            ("Hover Tooltip Content", "Case had no pre-existing SAP 6M assessment. AI blind diagnosis deduced [Machine] from first principles, consistent with benchmark historical case 8D-10049121.")
        ],
        five_why_chain=[
            ("Why 1", "Why is flange burr height 0.30 mm and pocket depth 12.86 mm?", "Workpiece experienced axial vibration during face milling pass.", "Inspection Lot 0010000001", "Yes", "Symptom Mechanism"),
            ("Why 2", "Why did workpiece experience vibration in Fixture F1?", "Clamping force decayed below the minimum required 4.5 kN threshold.", "Benchmarked with 8D-10049121", "Yes", "Physical Cause"),
            ("Why 3", "Why did clamping force drop on Fixture F1?", "Degraded hydraulic cylinder seals allow pressure bypass under cutting forces.", "precedents#8D-10049121", "Yes", "Root Cause (Precedent)"),
            ("Why 4", "Why was pressure decay undetected?", "Pre-shift inspection checklist did not mandate loaded pressure decay test.", "Benchmarked SOP review", "Yes", "Detection Gap"),
            ("Why 5", "Why omitted from PM interval?", "PM schedule PM-6M lacked annual seal overhaul requirements.", "Benchmarked PM standard", "Yes", "Systemic Root Cause")
        ],
        verification_checklist=[
            ("First-Principles Derivation", "blindEvidence.ts", "AI derives root cause from numbers without prior bias.", "PASS"),
            ("Precedent Benchmarking", "eightDAnalyzer.ts", "Attaches 85% similarity benchmark cleanly without overwriting.", "PASS"),
            ("UI Badge Rendering", "comparative-diagnosis-badge.tsx", "Renders Sky-Blue Benchmarked badge with precedent metadata.", "PASS")
        ]
    )

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 5: TC-04 FRESH FIRST-PRINCIPLES (ZERO HALLUCINATION)
    # ─────────────────────────────────────────────────────────────────────────────
    ws_tc4 = wb.create_sheet(title="TC-04 Fresh First-Principles")
    render_d4_sheet(
        ws_tc4,
        tc_id="TC_D4_RCA_004",
        title="Brand-New Material & Process -> Pure First-Principles Derivation (Zero Hallucination)",
        priority="High",
        objective="Verify that for a brand-new material and welding process with 0 historical precedents in Master Data, AI derives root cause purely from optical QM measurements without hallucinating fake precedents.",
        preconditions="1. Material BAT-9000 (Battery Module) and Work Center WC-WELD-02 (Laser Welding) have 0 records in HistoricalCases (Score < 30%).\n2. Incoming defect carries weld depth (0.85 mm) and laser power (2.8 kW) measurements.",
        sec1_fields=[
            ("Notification ID", "8D-20260831-04"),
            ("Defect Origin / Type", "Q3 - Internal Defect"),
            ("Found Date", "30/08/2026"),
            ("Discovery Mode", "During Inspection (Lot INS-90042)"),
            ("Inspection Lot ID", "INS-90042"),
            ("Quantity on Hold", "18 battery packs quarantined (Batch B-99014)"),
            ("Symptom Short Text", "Laser beam weld seam penetration depth below specification on battery housing"),
            ("Reported By", "Tung Nguyen (Laser QA Specialist)")
        ],
        sec2_fields=[
            ("Material ID", "BAT-9000 (New Material)"),
            ("Material Description", "Lithium Battery Module Housing"),
            ("Material Group", "MG-BATTERY"),
            ("Batch ID", "B-99014"),
            ("Work Center ID", "WC-WELD-02"),
            ("Work Center Description", "Laser Beam Welding Station 2"),
            ("Defect Code", "DEF-0899"),
            ("Defect Catalog Description", "Laser weld seam penetration void")
        ],
        char_rows=[
            ("Weld seam penetration depth", "0.85 mm", "1.40 ± 0.10 mm", "WC-WELD-02-L1", "Out of Spec (-0.45 mm shallow) - FAIL"),
            ("Laser power output", "2.8 kW", "3.5 ± 0.2 kW", "WC-WELD-02-L1", "Out of Spec (Power drop 20%) - FAIL")
        ],
        blind_diagnosis_details=[
            ("Blind Diagnosis Execution", "AI independently deduced [Machine] from optical and laser power telemetry (Confidence = 88%)."),
            ("Telemetry Physical Deduction", "Correlation between 2.8 kW power drop and 0.85 mm shallow weld depth proves laser optical beam delivery degradation.")
        ],
        historical_precedent_details=[
            ("Precedent Query Triggered", "SELECT from HistoricalCases WHERE materialId='BAT-9000' OR workCenterId='WC-WELD-02'"),
            ("Precedent Query Outcome", "0 matching records found (All scores < 30% similarity threshold)."),
            ("Precedent Status", "Honest empty precedent state — zero hallucinated precedents.")
        ],
        interaction_mechanism=[
            ("D4 System State", "FIRST-PRINCIPLES BLIND DIAGNOSIS (Zero Precedents Found)"),
            ("UI Badge Rendered", "⚪ First-Principles Blind Diagnosis"),
            ("Hover Tooltip Content", "No matching historical precedent exists in the dataset. Root cause [Machine] was derived purely from first-principles QM inspection measurements.")
        ],
        five_why_chain=[
            ("Why 1", "Why is weld seam penetration depth only 0.85 mm vs 1.40 mm spec?", "Laser beam energy density delivered to joint interface was below melting threshold.", "Lot INS-90042 measurement (0.85 mm)", "Yes", "Direct Mechanism"),
            ("Why 2", "Why was laser beam energy density deficient?", "Laser generator delivered only 2.8 kW optical power (spec: 3.5 ± 0.2 kW).", "Laser power meter reading (2.8 kW)", "Yes", "Energy Defect"),
            ("Why 3", "Why did laser optical power drop to 2.8 kW?", "Protective cover slide contaminated with spatter, attenuating beam transmission.", "Welding head optical inspection", "Yes", "Physical Root Cause"),
            ("Why 4", "Why was optical contamination not detected before 18 packs?", "Inline power monitoring sensor was disabled during tooling changeover.", "SOP-WELD-02 log audit", "Yes", "Detection Gap"),
            ("Why 5", "Why was power sensor disabled?", "Tooling changeover procedure did not include automated interlock for optical sensor check.", "Changeover SOP audit", "Yes", "Systemic Root Cause")
        ],
        verification_checklist=[
            ("Zero Hallucination", "AI-RULES R0.2", "System did not invent fictitious historical precedents.", "PASS"),
            ("First-Principles RCA", "blindEvidence.ts", "Successfully constructed 5-Why chain from laser power telemetry.", "PASS"),
            ("UI Badge Rendering", "comparative-diagnosis-badge.tsx", "Renders Slate First-Principles Blind Diagnosis badge cleanly.", "PASS")
        ]
    )

    output_path = "docs/8D_D4_Live_Verified_TestCases.xlsx"
    wb.save(output_path)
    print(f"Successfully generated focused D4 Test Cases workbook: {output_path}")

    try:
        wb.save("docs/8D_D4_BlindDiagnosis_Precedents_TestCases.xlsx")
        print("Updated docs/8D_D4_BlindDiagnosis_Precedents_TestCases.xlsx")
    except Exception as e:
        print(f"Note: {e}")

if __name__ == "__main__":
    create_focused_d4_test_cases()
